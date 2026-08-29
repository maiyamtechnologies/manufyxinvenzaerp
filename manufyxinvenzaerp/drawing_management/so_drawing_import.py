import frappe
from frappe import _
from frappe.utils import flt, now as frappe_now, generate_hash
from manufyxinvenzaerp.utils.dimension_formula import calculate_qty as _shared_calculate_qty


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _calc_qty(pig, length, width, thickness, unit_wt, sec_qty):
    """Return per-unit primary weight using the same formula as Drawing controller."""
    if pig in ("Structurals", "Plates"):
        qty = _shared_calculate_qty(pig, length, width, thickness, unit_wt, sec_qty)
        return qty if qty is not None else 0.0
    return flt(sec_qty)


def _get_file_path(file_url):
    file_doc = frappe.db.get_value(
        "File", {"file_url": file_url}, "name"
    )
    if not file_doc:
        frappe.throw(_("Attached file not found. Please re-attach."))
    return frappe.get_doc("File", file_doc).get_full_path()


def _parse_excel(file_path):
    """
    Parse BOM Excel and return (drawings_dict, raw_material_rows).

    drawings_dict  : OrderedDict  {cdn: {header fields, items: [...]}}
    raw_material_rows : list of dicts (flat, one per item row)
    """
    try:
        import openpyxl
    except ImportError:
        frappe.throw(_("openpyxl is required. Run: pip install openpyxl"))

    from collections import OrderedDict

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        frappe.throw(_("Could not open Excel file: {0}").format(str(e)))

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        frappe.throw(_("Excel file is empty."))

    header = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    col_idx = {h.lower(): i for i, h in enumerate(header)}

    def _get(row, *keys):
        for k in keys:
            i = col_idx.get(k.lower())
            if i is not None and i < len(row) and row[i] is not None:
                return row[i]
        return None

    def _sflt(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).strip().replace(",", ""))
        except Exception:
            return 0.0

    def _sstr(v):
        return str(v).strip() if v is not None else ""

    drawings = OrderedDict()

    for row in all_rows[1:]:
        cdn_raw = _get(row, "customer drawing number")
        if not cdn_raw:
            continue
        cdn = _sstr(cdn_raw)

        mat_raw = _get(row, "material code")
        if not mat_raw:
            continue
        mat_code = _sstr(mat_raw)

        if cdn not in drawings:
            fg_raw = _get(row, "fg item code", "fg item", "fg_item_code", "fg_item")
            drawings[cdn] = {
                "assembly_group": _sstr(_get(row, "assembly group")),
                "customer_drawing_number": cdn,
                "duno_mark_no": _sstr(_get(row, "duno/mark no", "duno mark no")),
                "fg_item_code": _sstr(fg_raw) if fg_raw else "",
                "total_quantity": _sflt(_get(row, "total qty")),
                "total_weight": _sflt(_get(row, "total weight (kg)", "total weight")),
                # Both are Drawing-level and are carried through as typed. They are
                # NOT validated here: the import stages rows with a raw SQL insert
                # that bypasses Link validation, so an unknown value has to reach the
                # staging table in order to be reported against its own row by
                # verify_raw_materials. Rejecting at parse time would abort the whole
                # file over one bad cell and say nothing about where it was.
                "nature_of_work": _sstr(_get(row, "nature of work", "nature_of_work")),
                "rate_schedule": _sstr(_get(row, "rate schedule", "rate_schedule")),
                "items": [],
            }

        drawings[cdn]["items"].append({
            "customer_drawing_number": cdn,
            "item_no": _sstr(_get(row, "item no")),
            "material_code": mat_code,
            "grade": _sstr(_get(row, "grade")),
            "thickness": _sflt(_get(row, "thickness")),
            "width": _sflt(_get(row, "width")),
            "length": _sflt(_get(row, "length")),
            "sec_qty": _sflt(_get(row, "reqd raw material qty", "reqd qty", "sec_qty")),
        })

    return drawings


# ---------------------------------------------------------------------------
# Whitelisted API
# ---------------------------------------------------------------------------

@frappe.whitelist()
def parse_bom_excel(so_name):
    """
    Parse the attached BOM Excel file on a Sales Order, then bulk-insert rows
    into Table 1 (Sales Order DUNO Item) and Table 2 (Sales Order Drawing Raw
    Material). Skips drawing numbers that already have a Drawing created.
    Returns {drawing_count, item_count, warnings, skipped_count}.
    """
    so = frappe.get_doc("Sales Order", so_name)
    if not so.get("custom_bom_excel_file"):
        frappe.throw(_("Please attach a BOM Excel file before loading."))

    file_path = _get_file_path(so.custom_bom_excel_file)
    drawings = _parse_excel(file_path)

    if not drawings:
        frappe.throw(_("No valid drawing rows found in the Excel file."))

    # Drawing numbers that already have a Drawing doc created — skip on reload
    locked_dnos = set(frappe.db.sql(
        """
        SELECT DISTINCT drawing_number
        FROM `tabSales Order DUNO Item`
        WHERE parent = %s AND drawing IS NOT NULL AND drawing != ''
        """,
        so_name,
        as_list=True,
    ))
    locked_dnos = {r[0] for r in locked_dnos}

    new_drawings = {cdn: d for cdn, d in drawings.items() if cdn not in locked_dnos}
    skipped_count = len(drawings) - len(new_drawings)

    if not new_drawings:
        return {
            "drawing_count": 0,
            "item_count": 0,
            "warnings": [_("{0} drawing(s) already have drawings created and were skipped.").format(skipped_count)],
            "skipped_count": skipped_count,
        }

    # --- Validate item codes (warn, don't throw) ---
    all_fg = {d["fg_item_code"] for d in new_drawings.values() if d["fg_item_code"]}
    all_mat = {i["material_code"] for d in new_drawings.values() for i in d["items"] if i["material_code"]}
    existing_items = set(frappe.db.get_all(
        "Item", filters={"name": ["in", list(all_fg | all_mat)]}, pluck="name"
    )) if (all_fg | all_mat) else set()

    warnings = []
    missing_fg = sorted(all_fg - existing_items)
    missing_mat = sorted(all_mat - existing_items)
    if missing_fg:
        warnings.append(_("FG Item codes not in Item master: {0}").format(", ".join(missing_fg)))
    if missing_mat:
        warnings.append(_("Material codes not in Item master: {0}").format(", ".join(missing_mat)))
    if skipped_count:
        warnings.append(_("{0} drawing(s) already created — skipped on reload.").format(skipped_count))

    # --- Fetch item master data for qty calculation ---
    item_data_map = {}
    if all_mat:
        for item in frappe.db.get_all(
            "Item",
            filters={"name": ["in", list(all_mat)]},
            fields=[
                "name", "item_name", "item_group", "custom_unit_weight",
                "custom_parent_item_group", "custom_secondary_uom", "stock_uom",
            ],
        ):
            item_data_map[item.name] = item

    # --- Warn for Plates missing thickness / Structurals missing unit weight in Excel ---
    # A dimension filled in for a group that does not use it is called out here
    # too, at the moment the sheet is read, so it reads against the sheet the
    # user is still looking at. verify_raw_materials repeats the check and is
    # the one that blocks -- this is the earlier, friendlier sighting.
    # Collected in the Raw Materials build loop below rather than here, because that
    # is the only place the row number these warnings land on is known -- t2_idx is
    # not calculated until further down, and it starts from whatever is already in
    # the table so a reload does not restart at 1.
    #
    # Naming the row is the whole point: "BEAM-1B16 -SHT-16 OF 291 / ISMB250" is the
    # drawing's name, not a place to go, and a sheet runs to hundreds of rows. Verify
    # Raw Materials already leads with the row (see _at); loading did not, so the two
    # messages described the same problem in two ways and only one of them could be
    # acted on.
    dim_warn = []

    # --- Fetch FG item names ---
    fg_name_map = {}
    if all_fg:
        for item in frappe.db.get_all(
            "Item",
            filters={"name": ["in", list(all_fg)]},
            fields=["name", "item_name"],
        ):
            fg_name_map[item.name] = item.item_name

    # --- Clear existing unlocked rows and reset verification flag ---
    frappe.db.sql(
        "DELETE FROM `tabSales Order Drawing Raw Material` WHERE parent = %s AND is_locked = 0",
        so_name,
    )
    frappe.db.sql(
        "DELETE FROM `tabSales Order DUNO Item` WHERE parent = %s AND (drawing IS NULL OR drawing = '')",
        so_name,
    )
    frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified", 0)

    # --- Re-index after delete: get next available idx for each table ---
    t1_max = frappe.db.sql(
        "SELECT COALESCE(MAX(idx), 0) FROM `tabSales Order DUNO Item` WHERE parent = %s", so_name
    )[0][0]
    t2_max = frappe.db.sql(
        "SELECT COALESCE(MAX(idx), 0) FROM `tabSales Order Drawing Raw Material` WHERE parent = %s", so_name
    )[0][0]
    t1_idx = int(t1_max) + 1
    t2_idx = int(t2_max) + 1

    now = frappe_now()
    user = frappe.session.user

    # --- Weight of the raw materials listed under each drawing ---
    # Filled at import so the Drawing List shows it beside the customer's own
    # figure straight after Load Items, without waiting for a save. Kept in step
    # afterwards by sales_order.recalculate_raw_material_qty, which recomputes it
    # from the rows on every save.
    calc_weight_by_cdn = {}
    for cdn, d in new_drawings.items():
        total = 0.0
        for item in d["items"]:
            idata = item_data_map.get(item["material_code"]) or frappe._dict()
            total += flt(_calc_qty(
                (idata.get("custom_parent_item_group") or "").strip(),
                item["length"], item["width"], item["thickness"],
                flt(idata.get("custom_unit_weight") or 0), flt(item["sec_qty"]),
            ))
        calc_weight_by_cdn[cdn] = flt(total, 3)

    # --- Build Table 1 (Drawing List) insert values ---
    t1_fields = [
        "name", "parent", "parenttype", "parentfield", "idx",
        "creation", "modified", "modified_by", "owner", "docstatus",
        "assembly_group", "item", "item_name", "duno_mark_no", "drawing_number",
        "total_quantity", "total_weight", "calculated_weight",
        "nature_of_work", "rate_schedule",
        "create_drawing", "submit_drawing", "mark_final_revision", "create_bom",
    ]
    t1_values = []
    for cdn, d in new_drawings.items():
        fg = d["fg_item_code"]
        t1_values.append((
            generate_hash(length=10),
            so_name, "Sales Order", "custom_duno_items", t1_idx,
            now, now, user, user, 0,
            d["assembly_group"], fg, fg_name_map.get(fg, ""),
            d["duno_mark_no"], cdn,
            d["total_quantity"], d["total_weight"], calc_weight_by_cdn.get(cdn, 0.0),
            d.get("nature_of_work") or "", d.get("rate_schedule") or "",
            1, 1, 1, 1,
        ))
        t1_idx += 1

    # --- Build Table 2 (Raw Materials) insert values ---
    t2_fields = [
        "name", "parent", "parenttype", "parentfield", "idx",
        "creation", "modified", "modified_by", "owner", "docstatus",
        "customer_drawing_number", "item_no", "material_code", "material_name",
        "item_group", "parent_item_group",
        "grade", "thickness", "width", "length",
        "sec_qty", "sec_uom", "total_sec_qty", "unit_weight", "qty", "uom", "total_weight", "is_locked",
    ]
    t2_values = []
    for cdn, d in new_drawings.items():
        tq = flt(d["total_quantity"]) or 1.0
        for item in d["items"]:
            idata = item_data_map.get(item["material_code"]) or frappe._dict()
            pig = (idata.get("custom_parent_item_group") or "").strip()
            unit_wt = flt(idata.get("custom_unit_weight") or 0)
            sec_qty = flt(item["sec_qty"])
            qty = _calc_qty(pig, item["length"], item["width"], item["thickness"], unit_wt, sec_qty)

            # Same three checks that used to run in their own loop above, now led by
            # the Raw Materials row they will occupy — the row Verify Raw Materials
            # will name for the same problem.
            _row_label = "%s / %s" % (cdn, item["material_code"])
            if pig == "Plates" and not flt(item["thickness"]):
                dim_warn.append(_at(RAW_MATERIALS, t2_idx,
                    _("{0}: Plates item missing Thickness in Excel").format(_row_label)))
            elif pig == "Structurals" and not unit_wt:
                dim_warn.append(_at(RAW_MATERIALS, t2_idx,
                    _("{0}: Structurals item missing Unit Weight in Item master").format(_row_label)))
            unused = _check_unused_dimensions(frappe._dict(item), pig)
            if unused:
                dim_warn.append(_at(RAW_MATERIALS, t2_idx,
                    _("{0}: {1} do not use {2} — clear that column in the sheet")
                    .format(_row_label, pig, ", ".join(unused))))

            total_sec_qty = flt(sec_qty * tq, 3)
            total_weight = flt(qty * tq, 3)
            t2_values.append((
                generate_hash(length=10),
                so_name, "Sales Order", "custom_so_raw_materials", t2_idx,
                now, now, user, user, 0,
                cdn,
                item["item_no"],
                item["material_code"],
                idata.get("item_name") or item["material_code"],
                idata.get("item_group") or "",
                pig,
                item["grade"],
                flt(item["thickness"], 3), flt(item["width"], 3), flt(item["length"], 3),
                flt(sec_qty, 3),
                idata.get("custom_secondary_uom") or "",
                flt(total_sec_qty, 3),
                flt(unit_wt, 6),
                flt(qty, 3),
                idata.get("stock_uom") or "",
                flt(total_weight, 3),
                0,
            ))
            t2_idx += 1

    if dim_warn:
        warnings.extend(dim_warn)

    # --- Bulk insert ---
    _bulk_insert("tabSales Order DUNO Item", t1_fields, t1_values)
    _bulk_insert("tabSales Order Drawing Raw Material", t2_fields, t2_values)
    frappe.db.commit()

    return {
        "drawing_count": len(t1_values),
        "item_count": len(t2_values),
        "warnings": warnings,
        "skipped_count": skipped_count,
    }


def _bulk_insert(table, fields, values, chunk_size=200):
    """Insert rows in chunks via raw SQL for performance with 1500+ rows."""
    if not values:
        return
    fields_sql = ", ".join([f"`{f}`" for f in fields])
    placeholders = "(" + ", ".join(["%s"] * len(fields)) + ")"

    for i in range(0, len(values), chunk_size):
        chunk = values[i : i + chunk_size]
        values_sql = ", ".join([placeholders] * len(chunk))
        flat = [v for row in chunk for v in row]
        frappe.db.sql(
            f"INSERT INTO `{table}` ({fields_sql}) VALUES {values_sql}",
            flat,
        )


# ---------------------------------------------------------------------------

@frappe.whitelist()
def create_drawings_from_import(so_name, batch_start=0, batch_size=30):
    """
    Create Draft Drawing documents in batches to avoid HTTP timeouts on large imports.
    Uses direct SQL to avoid loading the full SO doc (3000+ child rows) on every call.
    Returns {results, total, processed, next_start}.
    """
    batch_start = int(batch_start)
    batch_size = int(batch_size)

    # Lightweight SQL — never load the full SO doc with all child tables
    all_pending = frappe.db.sql(
        """SELECT name, drawing_number, item, duno_mark_no, total_quantity, total_weight
           , nature_of_work, rate_schedule
           FROM `tabSales Order DUNO Item`
           WHERE parent = %s AND create_drawing = 1
             AND (drawing IS NULL OR drawing = '')
           ORDER BY idx""",
        so_name, as_dict=True,
    )
    total = len(all_pending)
    if not all_pending:
        frappe.throw(_("No pending rows to create drawings for."))

    batch = all_pending[batch_start: batch_start + batch_size]
    if not batch:
        return {"results": [], "total": total, "processed": batch_start, "next_start": None}

    # SO header only — no child table load
    so_hdr = frappe.db.get_value(
        "Sales Order", so_name,
        ["customer", "customer_name", "project", "po_no"],
        as_dict=True,
    ) or frappe._dict()

    batch_cdns = [r.drawing_number for r in batch]

    # Load raw material rows only for this batch's CDNs
    rm_rows_raw = frappe.db.sql(
        """SELECT customer_drawing_number, item_no, material_code, material_name,
                  item_group, parent_item_group, grade, thickness, width, length,
                  sec_qty, sec_uom, unit_weight, uom
           FROM `tabSales Order Drawing Raw Material`
           WHERE parent = %s AND customer_drawing_number IN ({placeholders})
        """.format(placeholders=", ".join(["%s"] * len(batch_cdns))),
        [so_name] + batch_cdns, as_dict=True,
    )
    rm_by_cdn = {}
    for r in rm_rows_raw:
        rm_by_cdn.setdefault(r.customer_drawing_number, []).append(r)

    # Rate Schedule types for this batch, in one query.
    rs_names = list({r.rate_schedule for r in batch if r.get("rate_schedule")})
    rate_schedule_types = dict(frappe.get_all(
        "Rate Schedule", filters={"name": ["in", rs_names]},
        fields=["name", "type"], as_list=True,
    )) if rs_names else {}

    # Pre-fetch FG item data for this batch in one query
    fg_items = list({r.item for r in batch if r.item})
    item_cache = {}
    if fg_items:
        for item in frappe.db.get_all(
            "Item", filters={"name": ["in", fg_items]},
            fields=["name", "item_name", "description"],
        ):
            item_cache[item.name] = item

    results = []
    created = []  # (duno_row_name, drawing_name, cdn)

    for row_no, dr in enumerate(batch):
        cdn = dr.drawing_number
        result = {"drawing_number": cdn, "drawing": None, "status": "error", "error": ""}
        # One savepoint per drawing, so a failure undoes ONLY the drawing that
        # failed. Before this, the rollback in the handler below was a plain
        # frappe.db.rollback() -- it ended the whole transaction, taking every
        # drawing already inserted in this batch of 30 with it. The user saw one
        # error message and simply had fewer drawings than the sheet described,
        # with nothing saying which had been lost. On a 500-drawing import that is
        # up to 29 good drawings destroyed by one bad row.
        savepoint = "mfx_drawing_%d" % row_no
        frappe.db.savepoint(savepoint)
        try:
            item_data = item_cache.get(dr.item) or frappe._dict()
            no_of_qty = flt(dr.total_quantity) or 1

            drawing = frappe.get_doc({
                "doctype": "Drawing",
                "sales_order": so_name,
                "customer": so_hdr.customer,
                "customer_name": so_hdr.customer_name,
                "customer_no": so_hdr.customer,
                "project": so_hdr.get("project"),
                "cust_po_no": so_hdr.get("po_no"),
                "fg_item_code": dr.item or "",
                "fg_item_name": item_data.get("item_name") or "",
                "fg_description": item_data.get("description") or "",
                "no_of_qty_to_manufacture": flt(dr.total_quantity),
                "duno_mark_no": dr.duno_mark_no or "",
                "customer_drawing_number": cdn or "",
                "customer_provided_wt": flt(dr.total_weight),
                # Carried from the import sheet. verify_raw_materials has already
                # confirmed both exist in their masters, so these are safe to set as
                # Links here.
                "nature_of_work": dr.get("nature_of_work") or "",
                "rate_schedule": dr.get("rate_schedule") or "",
                # Type belongs to the Rate Schedule, so it is read from there rather
                # than carried in the sheet -- one less column to keep in step, and it
                # cannot contradict the schedule it describes. On the form Type is the
                # filter used to pick a schedule by hand; an import sets the schedule
                # directly, so without this it stayed blank and the picker then
                # filtered on an empty Type.
                "type": rate_schedule_types.get(dr.get("rate_schedule")) or "",
                "status": "Working",
            })

            for rm in rm_by_cdn.get(cdn, []):
                pig = rm.parent_item_group or ""
                unit_wt = flt(rm.unit_weight)
                sec_qty = flt(rm.sec_qty)
                qty = _calc_qty(pig, flt(rm.length), flt(rm.width), flt(rm.thickness), unit_wt, sec_qty)
                drawing.append("items", {
                    "item_number": rm.item_no or "",
                    "material_code": rm.material_code,
                    "material_name": rm.material_name or "",
                    "item_group": rm.item_group or "",
                    "parent_item_group": pig,
                    "thickness": flt(rm.thickness, 3),
                    "length": flt(rm.length, 3),
                    "width": flt(rm.width, 3),
                    "sec_qty": flt(sec_qty, 3),
                    "sec_uom": rm.sec_uom or "",
                    "unit_weight": flt(unit_wt, 6),
                    "qty": flt(qty, 3),
                    "uom": rm.uom or "",
                    "total_sec_qty": flt(sec_qty * no_of_qty, 3),
                    "total_qty": flt(qty * no_of_qty, 3),
                })

            drawing.insert(ignore_permissions=True)
            result["drawing"] = drawing.name
            result["status"] = "success"
            created.append((dr.name, drawing.name, cdn))

        except Exception as e:
            # Back to this drawing's own savepoint, not the start of the request.
            frappe.db.rollback(save_point=savepoint)
            frappe.local.message_log = []
            result["error"] = str(e)

        results.append(result)

    # Bulk-update links + locks once for the whole batch (not per-drawing)
    if created:
        for duno_name, drawing_name, _cdn in created:
            frappe.db.set_value(
                "Sales Order DUNO Item", duno_name, "drawing", drawing_name,
                update_modified=False,
            )
        cdns_created = [c[2] for c in created]
        frappe.db.sql(
            "UPDATE `tabSales Order Drawing Raw Material` SET is_locked = 1 "
            "WHERE parent = %s AND customer_drawing_number IN ({p})".format(
                p=", ".join(["%s"] * len(cdns_created))
            ),
            [so_name] + cdns_created,
        )
        frappe.db.commit()

    next_start = batch_start + len(batch)
    return {
        "results": results,
        "total": total,
        "processed": next_start,
        "next_start": next_start if next_start < total else None,
    }


# ---------------------------------------------------------------------------

@frappe.whitelist()
def process_drawings(so_name, step, batch_start=0, batch_size=30):
    """
    Run a single pipeline step in batches to avoid HTTP timeouts on 600+ drawings.
    Uses direct SQL to avoid loading the full SO doc on every call.
    Returns {results, total, processed, next_start}.
    """
    from manufyxinvenzaerp.drawing_management.drawing_utils import create_bom_from_drawing

    batch_start = int(batch_start)
    batch_size = int(batch_size)

    # Fetch qualifying DUNO rows via SQL (no SO doc load)
    all_rows = frappe.db.sql(
        """SELECT name, drawing, drawing_number,
                  submit_drawing, mark_final_revision, create_bom
           FROM `tabSales Order DUNO Item`
           WHERE parent = %s AND drawing IS NOT NULL AND drawing != ''
           ORDER BY idx""",
        so_name, as_dict=True,
    )
    total = len(all_rows)
    batch = all_rows[batch_start: batch_start + batch_size]

    if not batch:
        return {"results": [], "total": total, "processed": batch_start, "next_start": None}

    # Pre-fetch drawing metadata for the whole batch in ONE query
    drawing_names = [r.drawing for r in batch]
    drawing_meta = {
        d.name: d
        for d in frappe.db.get_all(
            "Drawing",
            filters={"name": ["in", drawing_names]},
            fields=["name", "docstatus", "status"],
        )
    }

    results = []

    for dr in batch:
        result = {
            "drawing": dr.drawing,
            "drawing_number": dr.drawing_number or "",
            "status": "skipped",
            "detail": "",
        }
        try:
            meta = drawing_meta.get(dr.drawing)
            if not meta:
                result["status"] = "error"
                result["error"] = "Drawing record not found"
                results.append(result)
                continue

            if step == "submit":
                if not dr.submit_drawing:
                    result["status"] = "unchecked"
                elif meta.docstatus != 0:
                    result["status"] = "already_done"
                else:
                    frappe.get_doc("Drawing", dr.drawing).submit()
                    result["status"] = "success"
                    result["detail"] = "submitted"

            elif step == "final_revision":
                if not dr.mark_final_revision:
                    result["status"] = "unchecked"
                elif meta.docstatus != 1:
                    result["status"] = "skipped"
                    result["detail"] = "not submitted"
                elif meta.status == "Final Revision":
                    result["status"] = "already_done"
                else:
                    # Direct set_value — no redundant full doc load
                    frappe.db.set_value("Drawing", dr.drawing, "status", "Final Revision")
                    result["status"] = "success"
                    result["detail"] = "marked final revision"

            elif step in ("create_bom", "create_and_submit_bom"):
                if not dr.create_bom:
                    result["status"] = "unchecked"
                elif meta.docstatus != 1 or meta.status != "Final Revision":
                    result["status"] = "skipped"
                    result["detail"] = "not in Final Revision"
                else:
                    bom_name = create_bom_from_drawing(dr.drawing)
                    if step == "create_and_submit_bom":
                        frappe.get_doc("BOM", bom_name).submit()
                        result["detail"] = "bom created and submitted: {0}".format(bom_name)
                    else:
                        result["detail"] = "bom: {0}".format(bom_name)
                    result["status"] = "success"

            elif step == "submit_bom":
                bom_name = frappe.db.get_value(
                    "BOM", {"custom_drawing": dr.drawing, "docstatus": 0}, "name"
                )
                if not bom_name:
                    result["status"] = "skipped"
                    result["detail"] = "no draft BOM"
                else:
                    frappe.get_doc("BOM", bom_name).submit()
                    result["status"] = "success"
                    result["detail"] = "bom submitted: {0}".format(bom_name)

            else:
                frappe.throw(_("Unknown step: {0}").format(step))

        except Exception as e:
            frappe.db.rollback()
            frappe.local.message_log = []
            result["status"] = "error"
            result["error"] = str(e)

        results.append(result)

    frappe.db.commit()

    next_start = batch_start + len(batch)
    return {
        "results": results,
        "total": total,
        "processed": next_start,
        "next_start": next_start if next_start < total else None,
    }


# ---------------------------------------------------------------------------

def _at(table, idx, text):
    """Put the row number in front of the complaint.

    A verification run reports against a sheet of hundreds of rows, and "Drawing
    BEAM-1B16 -SHT-16 OF 291 / ISMB250 (Item 1w11)" is the drawing's name, not a place
    to go. The row number is, so it leads.

    The table is named with it because the Sales Order has two the issues can come
    from -- Drawing List on the Drawing Import tab, and Raw Materials -- and row 16 of
    one is not row 16 of the other.
    """
    return "<b>%s</b> &middot; %s" % (_("{0} row {1}").format(_(table), idx), text)


# Plain strings, translated inside _at: calling _() at import time runs outside any
# request and would pin the message to whatever language the worker started in.
DRAWING_LIST = "Drawing List"
RAW_MATERIALS = "Raw Materials"


def _check_drawing_masters(so):
    """Nature of Work / Rate Schedule on each imported drawing row must already exist
    in their masters.

    Validated by record NAME and nothing else -- no format rule. Rate Schedule is named
    by its own RS No, so the name IS the title being typed ("RS- O/S-001 A"), and a
    pattern invented from one example would start rejecting valid codes the moment the
    client's numbering changed. Existence cannot go stale that way.

    Checked here rather than at parse time because the import stages rows with a raw
    SQL insert, which bypasses Link validation: a wrong value has to land in the table
    to be reported against the drawing it came from. Anything reported blocks
    verification, so the file has to be corrected before the flow can go on.

    Blank is allowed -- neither is mandatory on a Drawing, and older imports predate
    both columns."""
    rows = [r for r in (so.get("custom_duno_items") or [])]
    if not rows:
        return []

    wanted = {"Nature of Work": set(), "Rate Schedule": set()}
    for r in rows:
        if r.get("nature_of_work"):
            wanted["Nature of Work"].add(r.nature_of_work)
        if r.get("rate_schedule"):
            wanted["Rate Schedule"].add(r.rate_schedule)

    existing = {
        doctype: set(frappe.get_all(doctype, filters={"name": ["in", list(names)]}, pluck="name"))
        if names else set()
        for doctype, names in wanted.items()
    }

    issues = []
    for r in rows:
        label = r.get("drawing_number") or r.get("duno_mark_no") or "?"
        for doctype, fieldname in (("Nature of Work", "nature_of_work"),
                                   ("Rate Schedule", "rate_schedule")):
            value = r.get(fieldname)
            if value and value not in existing[doctype]:
                issues.append(_at(DRAWING_LIST, r.idx,
                    _("Drawing {0}: {1} <b>{2}</b> is not in the {1} master. "
                      "Correct it in the sheet (or create the record) and import again.")
                    .format(label, doctype, value)))
    return issues


# Every dimension the weight formula reads, per group. A dimension NOT listed
# for a group takes no part in that group's formula, so a value sitting in it
# describes nothing -- see _check_unused_dimensions.
GROUP_DIMENSIONS = {
    "Structurals": {"length"},
    "Plates": {"length", "width", "thickness"},
    "Nuts and Bolts": set(),
}
DIMENSION_SHEET_COLUMNS = {
    "length": "Length",
    "width": "Width",
    "thickness": "Thickness",
}


def _check_row_required(row, group):
    """Inputs the group's formula cannot produce a weight without.

    Mirrors dimension_formula.calculate_qty exactly: Structurals need Length x
    Unit Weight x Sec Qty, Plates need those plus Width and Thickness, and Nuts
    and Bolts convert Sec Qty (Nos) with the Unit Weight. Sec Qty was missing
    from this list before -- a blank "Reqd Raw Material Qty" column made the
    formula return nothing and the row was staged weighing zero, silently.
    """
    missing = []
    for dim in sorted(GROUP_DIMENSIONS.get(group) or ()):
        if not flt(row.get(dim)):
            missing.append(DIMENSION_SHEET_COLUMNS[dim])
    if not flt(row.get("sec_qty")):
        missing.append(_("Reqd Raw Material Qty"))
    if not flt(row.get("unit_weight")):
        missing.append(_("Unit Weight (set it on the Item master)"))
    return missing


def _check_unused_dimensions(row, group):
    """Dimensions filled in for a group whose formula never reads them.

    A Structural's weight is Length x Unit Weight x Sec Qty -- Thickness and
    Width play no part, so a Thickness typed against a beam is not a harmless
    extra: it is copied into the Drawing, the BOM and Material Planning as a
    real requirement, and Purchase Receipt matches batches on all three
    dimensions strictly. The beam then arrives with no thickness, fails to
    match, and the whole received batch is routed to Material Mapping instead
    of Exact Match. Caught here, at the one point before drawings exist.
    """
    used = GROUP_DIMENSIONS.get(group)
    if used is None:
        return []
    return [
        DIMENSION_SHEET_COLUMNS[dim]
        for dim in sorted(set(DIMENSION_SHEET_COLUMNS) - used)
        if flt(row.get(dim))
    ]


def _check_drawing_headers(so):
    """Header columns of each drawing that has not been created yet.

    A blank FG Item or Total Qty is not reported anywhere else: the import
    substitutes 1.0 for a missing Total Qty, so every total on the drawing
    would silently be computed for a single piece.
    """
    rows = [r for r in (so.get("custom_duno_items") or []) if not r.get("drawing")]
    if not rows:
        return []

    fg_codes = {r.item for r in rows if r.get("item")}
    existing_fg = set(frappe.get_all(
        "Item", filters={"name": ["in", list(fg_codes)]}, pluck="name"
    )) if fg_codes else set()

    issues = []
    for r in rows:
        label = r.get("drawing_number") or r.get("duno_mark_no") or "?"
        if not r.get("item"):
            issues.append(_at(DRAWING_LIST, r.idx,
                              _("Drawing {0}: FG Item is missing").format(label)))
        elif r.item not in existing_fg:
            issues.append(_at(DRAWING_LIST, r.idx,
                              _("Drawing {0}: FG Item <b>{1}</b> is not in the Item master")
                              .format(label, r.item)))
        if not r.get("duno_mark_no"):
            issues.append(_at(DRAWING_LIST, r.idx,
                              _("Drawing {0}: DUNO/Mark No is missing").format(label)))
        if flt(r.get("total_quantity")) <= 0:
            issues.append(_at(DRAWING_LIST, r.idx,
                _("Drawing {0}: Total Qty is {1} — set how many are being made, "
                  "otherwise every total is calculated for one piece.")
                .format(label, flt(r.get("total_quantity")))))
    return issues


@frappe.whitelist()
def verify_raw_materials(so_name):
    """
    Validate all unlocked Raw Material rows on the Sales Order.
    Sets custom_raw_materials_verified = 1 if no issues found.
    Returns {issues: [...], verified: bool}.

    This is the only gate between the uploaded sheet and the documents built
    from it, so it checks every value the weight formula reads -- present ones
    that should not be, absent ones that must be, and the weight the row ends
    up carrying -- rather than a subset. Anything it reports has to be fixed in
    the sheet and the file re-loaded; nothing here edits the staged rows,
    because the sheet stays the single source of truth.
    """
    so = frappe.get_doc("Sales Order", so_name)
    # r.get("is_locked"), never r.is_locked: frappe's Document class defines
    # is_locked as a property (whether a FILE LOCK is held on the document),
    # and a class property shadows the field of the same name on every
    # instance -- the attribute reads False whatever the column holds. Reading
    # it that way made this list every row, so rows already turned into
    # Drawings were re-verified and could fail on a sheet correction that no
    # longer applies to them.
    unlocked = [r for r in (so.custom_so_raw_materials or []) if not r.get("is_locked")]
    issues = _check_drawing_masters(so) + _check_drawing_headers(so)

    if not unlocked:
        verified = not issues
        frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified",
                            1 if verified else 0, update_modified=False)
        frappe.db.commit()
        modified = frappe.db.get_value("Sales Order", so_name, "modified")
        return {"issues": issues, "verified": verified, "modified": str(modified)}

    all_mat = {r.material_code for r in unlocked if r.material_code}
    # The group and unit weight are read from the Item master rather than the
    # staged row: the row holds a copy taken at import time, and an Item edited
    # afterwards would otherwise be verified against values no longer in force.
    item_master = {
        d.name: d
        for d in frappe.get_all(
            "Item",
            filters={"name": ["in", list(all_mat)]},
            fields=["name", "custom_parent_item_group", "custom_unit_weight"],
        )
    } if all_mat else {}

    # Drawings still carrying rows are reported once if they have none at all.
    seen_cdns = {r.customer_drawing_number for r in unlocked if r.customer_drawing_number}
    for r in (so.get("custom_duno_items") or []):
        if not r.get("drawing") and r.get("drawing_number") and r.drawing_number not in seen_cdns:
            issues.append(_at(DRAWING_LIST, r.idx,
                              _("Drawing {0}: no raw material rows were loaded for it")
                              .format(r.drawing_number)))

    for row in unlocked:
        cdn = row.customer_drawing_number or "?"
        mat = row.material_code or ""
        item_no = row.item_no or "?"

        def _issue(text, _row=row):
            issues.append(_at(RAW_MATERIALS, _row.idx,
                              _("Drawing {0} / {1} (Item {2}): {3}")
                              .format(cdn, mat or "?", item_no, text)))

        if not mat:
            issues.append(_at(RAW_MATERIALS, row.idx,
                              _("Drawing {0}, Item {1}: Material Code is missing")
                              .format(cdn, item_no)))
            continue

        idata = item_master.get(mat)
        if not idata:
            issues.append(_at(RAW_MATERIALS, row.idx,
                              _("Drawing {0} / {1}: Not found in Item master")
                              .format(cdn, mat)))
            continue

        group = (idata.custom_parent_item_group or "").strip()
        if group not in GROUP_DIMENSIONS:
            _issue(_("Parent Item Group on the Item master is <b>{0}</b> — it must be one of {1}, "
                     "or no weight can be calculated for this row.")
                   .format(group or _("not set"), ", ".join(sorted(GROUP_DIMENSIONS))))
            continue

        if (row.parent_item_group or "").strip() != group:
            _issue(_("staged as <b>{0}</b> but the Item master now says <b>{1}</b> — "
                     "re-load the sheet so the rows match the master.")
                   .format(row.parent_item_group or _("blank"), group))
            continue

        missing = _check_row_required(row, group)
        if missing:
            _issue(_("{0} — missing {1}").format(group, ", ".join(missing)))

        unused = _check_unused_dimensions(row, group)
        if unused:
            _issue(_("{0} do not use {1} — clear that column in the sheet. "
                     "A value there is carried into the Drawing, BOM and Material Planning "
                     "as a real requirement that the delivered material can never match.")
                   .format(group, ", ".join(unused)))

        # Last line of defence: whatever the reason, a row that ends up weighing
        # nothing must not reach a Drawing.
        expected = flt(_calc_qty(group, row.length, row.width, row.thickness,
                                 flt(idata.custom_unit_weight), row.sec_qty), 3)
        if expected <= 0:
            if not missing:
                _issue(_("calculated weight is zero — check the dimensions and the Item master's Unit Weight."))
        elif abs(expected - flt(row.qty, 3)) > 0.01:
            _issue(_("row weighs {0} Kg but the current dimensions and Unit Weight give {1} Kg — "
                     "re-load the sheet.").format(flt(row.qty, 3), expected))

    verified = len(issues) == 0
    frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified", 1 if verified else 0, update_modified=False)
    frappe.db.commit()
    modified = frappe.db.get_value("Sales Order", so_name, "modified")
    return {"issues": issues, "verified": verified, "modified": str(modified)}


@frappe.whitelist()
def download_bom_template():
    """Return a pre-filled BOM Import Excel template as a file download."""
    try:
        import openpyxl
    except ImportError:
        frappe.throw(_("openpyxl is required. Run: pip install openpyxl"))

    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM Import"

    # Nature of Work and Rate Schedule are Drawing-level, so they sit with the other
    # header columns and repeat on every row of a drawing, same as Assembly Group.
    # Both must be the master record's NAME exactly as it reads there -- Rate Schedule
    # is named by its own RS No, so what the client types IS the record.
    headers = [
        "Assembly Group", "Customer Drawing Number", "DUNO/Mark No",
        "FG Item", "Total Qty", "Total Weight (KG)",
        "Nature of Work", "Rate Schedule",
        "Item No", "Material Code", "Grade", "Thickness", "Width", "Length",
        "Reqd Raw Material Qty",
    ]
    ws.append(headers)

    # Samples use whatever is really in the masters, so the template a client
    # downloads is filled in with values that will actually verify rather than
    # placeholders they have to guess the shape of.
    sample_now = frappe.db.get_value("Nature of Work", {}, "name") or "Auto Welding"
    sample_rs = frappe.db.get_value("Rate Schedule", {}, "name") or "RS-001"

    # Sample row 1 — drawing CDN-001, item 1
    ws.append([
        "Structural Assembly", "CDN-001", "DM-001", "FG-ITEM-001", 5, 250.0,
        sample_now, sample_rs,
        "1", "MAT-STRUCT-001", "A36", 0, 0, 3000, 2,
    ])
    # Sample row 2 — same drawing CDN-001, item 2 (same header columns repeated)
    ws.append([
        "Structural Assembly", "CDN-001", "DM-001", "FG-ITEM-001", 5, 250.0,
        sample_now, sample_rs,
        "2", "MAT-PLATE-001", "IS2062", 10, 200, 1500, 1,
    ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = "BOM_Import_Template.xlsx"
    frappe.response["filecontent"] = output.read()
    frappe.response["type"] = "download"


@frappe.whitelist()
def clear_drawing_import(so_name):
    """
    Remove the BOM Excel attachment and delete all unlocked import rows.
    Rows that already have a Drawing created (locked) are preserved.
    Returns counts of deleted rows.
    """
    # Count what will be removed
    t1_del = frappe.db.sql(
        "SELECT COUNT(*) FROM `tabSales Order DUNO Item` WHERE parent = %s AND (drawing IS NULL OR drawing = '')",
        so_name, as_list=True
    )[0][0]
    t2_del = frappe.db.sql(
        "SELECT COUNT(*) FROM `tabSales Order Drawing Raw Material` WHERE parent = %s AND is_locked = 0",
        so_name, as_list=True
    )[0][0]

    # Delete unlocked rows
    frappe.db.sql(
        "DELETE FROM `tabSales Order DUNO Item` WHERE parent = %s AND (drawing IS NULL OR drawing = '')",
        so_name,
    )
    frappe.db.sql(
        "DELETE FROM `tabSales Order Drawing Raw Material` WHERE parent = %s AND is_locked = 0",
        so_name,
    )

    # Clear the file attachment field and verification flag on the SO
    frappe.db.set_value("Sales Order", so_name, "custom_bom_excel_file", "")
    frappe.db.set_value("Sales Order", so_name, "custom_raw_materials_verified", 0)
    frappe.db.commit()

    return {"deleted_drawings": int(t1_del), "deleted_items": int(t2_del)}


@frappe.whitelist()
def get_cancelled_drawing_links(sales_order):
    """DUNO rows on this order still pointing at a cancelled drawing.

    Read by the Sales Order form so the problem is named on screen -- which DUNO, and
    what to do -- rather than discovered as "Cannot link cancelled document: Row #22"
    when somebody tries to save or submit.

    It has to be the form that says it. Frappe checks links in _validate_links(),
    which runs before every server-side hook this app could use, so a message raised
    from validate() or before_submit() is never reached.

    Cancelling a drawing now releases its row by itself, so this is for orders that
    were already in that state when the release was added."""
    rows = frappe.get_all(
        "Sales Order DUNO Item",
        filters={"parent": sales_order, "drawing": ["!=", ""]},
        fields=["idx", "duno_mark_no", "drawing_number", "drawing"],
        order_by="idx asc",
    )
    if not rows:
        return []

    cancelled = set(frappe.get_all(
        "Drawing",
        filters={"name": ["in", [r.drawing for r in rows]], "docstatus": 2},
        pluck="name",
    ))
    return [r for r in rows if r.drawing in cancelled]
